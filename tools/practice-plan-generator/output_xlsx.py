"""
output_xlsx.py — Generate a .xlsx practice plan (v2 — 3-phase structure)
Sheet 1: Practice Plan  |  Sheet 2: Skills Coverage
"""
from pathlib import Path
from datetime import datetime

OUTPUT_DIR = Path(__file__).parent / "outputs"

NAVY     = "1B3A6B"
RED      = "C8102E"
LT_BLUE  = "E8EFF8"
ALT_ROW  = "F5F7FA"
WHITE    = "FFFFFF"
DARK     = "1A1A2E"
GRAY     = "888888"
MED_BLUE = "2E5FA3"
GREEN_BG = "E8F5E9"
AMBER_BG = "FFF8E1"
DARK_GREEN = "1B5E20"
DARK_AMBER = "E65100"

PHASE_HEADER_COLORS = {
    "opening":  (LT_BLUE,  NAVY),
    "stations": (GREEN_BG, DARK_GREEN),
    "team":     (AMBER_BG, DARK_AMBER),
    "closure":  (LT_BLUE,  NAVY),
}

GAME_LIKE_TEAM_IDS = {
    "two_pitch_game", "game_situation_scrimmage", "live_bp_rotation",
    "live_bp_wiffle_rotation", "fielding_doubles_game", "fielding_singles_game",
    "infield_outfield_combined",
}


def generate_xlsx(plan: dict) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Practice Plan"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 10   # Time / Station label
    ws.column_dimensions["B"].width = 26   # Drill/Activity name
    ws.column_dimensions["C"].width = 52   # Description
    ws.column_dimensions["D"].width = 38   # Coaching cues
    ws.column_dimensions["E"].width = 8    # Level

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(name="Arial", size=10, bold=False, italic=False, color="000000"):
        return Font(name=name, size=size, bold=bold, italic=italic, color=color)

    def align(h="left", v="top", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def write(ws, row, col, value, bg=WHITE, fg=DARK, sz=10, bold=False,
              italic=False, h="left", v="top", wrap=True, height=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = fill(bg)
        cell.font = font(size=sz, bold=bold, italic=italic, color=fg)
        cell.alignment = align(h, v, wrap)
        if height:
            ws.row_dimensions[row].height = height
        return cell

    row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    write(ws, row, 1,
          "LGLL PRACTICE PLAN  ·  La Grange Little League  ·  Ages 7–9",
          bg=NAVY, fg=WHITE, sz=13, bold=True, h="center", v="center", height=26)
    row += 1

    # ── Meta ──────────────────────────────────────────────────────────────────
    meta_labels = ["DATE", "LOCATION", "DURATION / COACHES", "FOCUS", ""]
    meta_values = [
        plan["date"], plan["location"],
        f"{plan['duration']} min  ·  {plan['coaches']} coaches  ·  ~{plan['player_count']} players",
        plan["focus"], "",
    ]
    for ci, lbl in enumerate(meta_labels, 1):
        col = get_column_letter(ci)
        write(ws, row,   ci, lbl, bg=MED_BLUE, fg=WHITE, sz=8, bold=True, h="center", v="center", wrap=False)
        write(ws, row+1, ci, meta_values[ci-1], bg=LT_BLUE, fg=DARK, sz=9, h="center", v="center", wrap=False)
    ws.row_dimensions[row].height   = 13
    ws.row_dimensions[row+1].height = 16
    row += 3

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["TIME", "DRILL / ACTIVITY", "DESCRIPTION", "COACHING CUES", "LEVEL"]
    for ci, h in enumerate(headers, 1):
        write(ws, row, ci, h, bg=NAVY, fg=WHITE, sz=9, bold=True, h="center", v="center", wrap=False)
    ws.row_dimensions[row].height = 15
    row += 1

    # ── Phases ────────────────────────────────────────────────────────────────
    alt = False

    def drill_row(drill, minutes, bg, sub_label=None):
        nonlocal row
        d = drill
        time_label = f"{minutes} min"

        write(ws, row, 1, time_label, bg=bg, fg=RED, sz=10, bold=True, h="center", v="top")
        name = d["name"]
        if d.get("level") == "intermediate":
            name += "  ★"
        if sub_label:
            name = f"[{sub_label}]\n{name}"
        write(ws, row, 2, name, bg=bg, fg=DARK, sz=10, bold=True, v="top")

        desc = d.get("description","").replace("\n"," ").strip()
        write(ws, row, 3, desc, bg=bg, fg="333333", sz=9, v="top")

        cues = d.get("coaching_cues", [])
        write(ws, row, 4, "\n".join(f"▸ {c}" for c in cues),
              bg=bg, fg=NAVY, sz=8, italic=True, v="top")

        write(ws, row, 5, d.get("level","basic").upper(),
              bg=bg, fg=(RED if d.get("level")=="intermediate" else GRAY),
              sz=8, bold=(d.get("level")=="intermediate"), h="center", v="top")

        line_est = max(2, len(cues)) * 14
        ws.row_dimensions[row].height = line_est
        row += 1

    for phase in plan["phases"]:
        bg_h, fg_h = PHASE_HEADER_COLORS.get(phase["type"], (LT_BLUE, NAVY))
        t   = phase["start"]
        end = t + phase["minutes"]

        # Phase header
        ws.merge_cells(f"A{row}:E{row}")
        header_text = f"  {phase['label'].upper()}  ·  {phase['minutes']} min  ·  T+{t:02d}–{end:02d}"
        if phase["type"] == "stations":
            sd = phase["data"]
            header_text += f"  ·  {sd['n_rotations']} stations × {sd['rotation_minutes']} min — rotate on whistle"
        write(ws, row, 1, header_text, bg=bg_h, fg=fg_h, sz=11, bold=True,
              h="left", v="center", wrap=False, height=18)
        row += 1

        if phase["type"] == "opening":
            cur_sub = None
            for item in phase["items"]:
                bg = ALT_ROW if alt else WHITE
                alt = not alt
                drill_row(item["drill"], item["minutes"], bg,
                          sub_label=item["sub_phase"] if item["sub_phase"] != cur_sub else None)
                cur_sub = item["sub_phase"]

        elif phase["type"] == "stations":
            sd = phase["data"]
            ws.merge_cells(f"A{row}:E{row}")
            write(ws, row, 1,
                  f"  Groups rotate every {sd['rotation_minutes']} minutes. "
                  f"Each player visits all {sd['n_rotations']} stations.",
                  bg=WHITE, fg=GRAY, sz=9, italic=True, h="left", v="center",
                  wrap=False, height=14)
            row += 1

            letters = "ABCDE"
            for i, s in enumerate(sd["stations"]):
                d = s["drill"]
                coach_note = "coach present" if s["coach_assigned"] else "self-directed"

                # Station sub-header
                ws.merge_cells(f"A{row}:E{row}")
                write(ws, row, 1,
                      f"  Station {letters[i]} · {s['label']}  [{coach_note} · ~{s['group_size']} players · {s['minutes']} min]",
                      bg=MED_BLUE, fg=WHITE, sz=9, bold=True, h="left", v="center",
                      wrap=False, height=16)
                row += 1

                bg = ALT_ROW if alt else WHITE
                alt = not alt
                drill_row(d, s["minutes"], bg)

        elif phase["type"] == "team":
            for item in phase["items"]:
                d   = item["drill"]
                tag = "GAME-LIKE" if d["id"] in GAME_LIKE_TEAM_IDS else "INSTRUCTIVE"
                bg  = ALT_ROW if alt else WHITE
                alt = not alt

                # Tag badge in time cell
                cell_a = ws.cell(row=row, column=1, value=f"{item['minutes']} min\n{tag}")
                cell_a.fill = fill(bg)
                tag_color = DARK_GREEN if tag == "GAME-LIKE" else DARK_AMBER
                cell_a.font = Font(name="Arial", size=9, bold=True, color=tag_color)
                cell_a.alignment = align("center","top")

                name = d["name"]
                if d.get("level") == "intermediate":
                    name += "  ★"
                write(ws, row, 2, name, bg=bg, fg=DARK, sz=11, bold=True, v="top")
                desc = d.get("description","").replace("\n"," ").strip()
                write(ws, row, 3, desc, bg=bg, fg="333333", sz=9, v="top")
                cues = d.get("coaching_cues",[])
                write(ws, row, 4, "\n".join(f"▸ {c}" for c in cues),
                      bg=bg, fg=NAVY, sz=8, italic=True, v="top")
                write(ws, row, 5, d.get("level","basic").upper(),
                      bg=bg, fg=(RED if d.get("level")=="intermediate" else GRAY),
                      sz=8, h="center", v="top")
                ws.row_dimensions[row].height = max(2, len(cues)) * 15
                row += 1

        elif phase["type"] == "closure":
            bg = LT_BLUE
            write(ws, row, 1, "5 min", bg=bg, fg=RED, sz=10, bold=True, h="center", v="top")
            write(ws, row, 2, "Team Huddle & Closure", bg=bg, fg=DARK, sz=10, bold=True, v="top")
            write(ws, row, 3,
                  "2-3 specific positives. One focus area for next practice. Team cheer.",
                  bg=bg, fg="333333", sz=9, v="top")
            write(ws, row, 4, "", bg=bg)
            write(ws, row, 5, "BASIC", bg=bg, fg=GRAY, sz=8, h="center", v="top")
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1  # blank row between phases

    # ── Equipment ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    write(ws, row, 1, "  EQUIPMENT NEEDED", bg=LT_BLUE, fg=NAVY, sz=10, bold=True,
          h="left", v="center", wrap=False, height=16)
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    eq_items = [e.replace("_"," ").title() for e in plan["all_equipment"]]
    write(ws, row, 1, "  " + "  ·  ".join(eq_items) if eq_items else "  No special equipment",
          bg=WHITE, fg=DARK, sz=9, h="left", v="center", wrap=True, height=18)
    row += 2

    # ── Footer ────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    write(ws, row, 1,
          f"Generated {datetime.now().strftime('%B %d, %Y')}  ·  "
          f"LGLL Practice Plan Generator  ·  {plan['total_minutes']} / {plan['duration']} min",
          bg=WHITE, fg=GRAY, sz=7, h="center", v="center", wrap=False, height=12)

    # ── Sheet 2: Skills Coverage ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Skills Coverage")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12

    ws2.merge_cells("A1:B1")
    write(ws2, 1, 1, "SKILLS COVERED THIS PRACTICE",
          bg=NAVY, fg=WHITE, sz=12, bold=True, h="center", v="center",
          wrap=False, height=22)

    write(ws2, 2, 1, "SKILL", bg=MED_BLUE, fg=WHITE, sz=9, bold=True,
          h="center", v="center", wrap=False)
    write(ws2, 2, 2, "COVERED", bg=MED_BLUE, fg=WHITE, sz=9, bold=True,
          h="center", v="center", wrap=False)
    ws2.row_dimensions[2].height = 13

    covered = set(plan.get("all_skills_covered", []))
    # Collect all skills from all drills in all phases
    if not covered:
        for phase in plan["phases"]:
            items = []
            if phase["type"] in ("opening", "team", "closure"):
                items = phase.get("items", [])
            elif phase["type"] == "stations":
                items = [{"drill": s["drill"]} for s in phase["data"]["stations"]]
            for item in items:
                for sk in item["drill"].get("skills", []):
                    covered.add(sk)

    import yaml
    with open(Path(__file__).parent / "data" / "rubric_skills.yaml") as f:
        rubric = yaml.safe_load(f)

    r = 3
    for cat_key, cat_data in rubric["categories"].items():
        ws2.merge_cells(f"A{r}:B{r}")
        write(ws2, r, 1, cat_data["label"].upper(),
              bg=LT_BLUE, fg=NAVY, sz=9, bold=True, h="left", v="center",
              wrap=False, height=16)
        r += 1

        for skill_key, skill_data in cat_data["skills"].items():
            is_covered = skill_key in covered
            bg_c = "E8F5E9" if is_covered else WHITE

            write(ws2, r, 1, "  " + skill_data["label"],
                  bg=bg_c, fg=DARK, sz=9, h="left", v="center", wrap=False)
            write(ws2, r, 2, "✓" if is_covered else "",
                  bg=bg_c, fg=("2E7D32" if is_covered else GRAY),
                  sz=10, bold=is_covered, h="center", v="center", wrap=False)
            ws2.row_dimensions[r].height = 14
            r += 1

            if "sub_skills" in skill_data:
                for sub in skill_data["sub_skills"]:
                    is_sub = sub in covered
                    bg_s = "F1F8F1" if is_sub else "FAFAFA"
                    write(ws2, r, 1, "    › " + sub.replace("_"," ").title(),
                          bg=bg_s, fg=GRAY, sz=8, italic=True, h="left", v="center", wrap=False)
                    write(ws2, r, 2, "✓" if is_sub else "",
                          bg=bg_s, fg=("2E7D32" if is_sub else GRAY),
                          sz=9, bold=is_sub, h="center", v="center", wrap=False)
                    ws2.row_dimensions[r].height = 12
                    r += 1

    # ── Save ──────────────────────────────────────────────────────────────────
    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"practice_plan_{date_str}_{plan['location_key']}_{plan['duration']}min.xlsx"
    out_path = OUTPUT_DIR / filename
    wb.save(str(out_path))
    return str(out_path)
